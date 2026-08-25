import json
import time
import re
import glob
import zipfile
import tempfile
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from typing import Dict, Optional
from datetime import datetime
import sys
import os

def resource_path(filename):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.abspath("."), filename)

def get_service_account_path():
    appdata = os.getenv('APPDATA')
    if not appdata:
        return resource_path("service-account.json")
    
    app_folder = os.path.join(appdata, "UbuySEOAutomation")
    if not os.path.exists(app_folder):
        os.makedirs(app_folder, exist_ok=True)
        
    appdata_path = os.path.join(app_folder, "service-account.json")
    
    # If the file does not exist in AppData, copy it from the bundled resource
    if not os.path.exists(appdata_path):
        import shutil
        bundled_path = resource_path("service-account.json")
        if os.path.exists(bundled_path) and bundled_path != appdata_path:
            try:
                shutil.copy(bundled_path, appdata_path)
                print(f"Copied service-account.json from bundled resources to {appdata_path}")
            except Exception as e:
                print(f"Failed to copy service-account.json: {e}")
                
    # Fallback to current directory or bundled resource if still not found
    if not os.path.exists(appdata_path):
        return resource_path("service-account.json")
        
    return appdata_path

def normalize_number_string(s: str) -> int:
    s = s.strip()
    if not s:
        return 0
    m = re.match(r"([\d.,]+)\s*([KMkm])?", s)
    if not m:
        return 0
    num = m.group(1).replace(",", "")
    suffix = m.group(2).upper() if m.group(2) else ""
    try:
        val = float(num)
    except ValueError:
        return 0
    if suffix == "K":
        val *= 1000
    elif suffix == "M":
        val *= 1000000
    return int(round(val))


def get_index_counts(domain: str, driver) -> Dict[str, int]:
    """Fetch indexed and non-indexed URL counts from GSC."""
    target = f"https://search.google.com/search-console/index?resource_id=sc-domain:{domain}"
    driver.get(target)
    wait = WebDriverWait(driver, 25)
    wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.nnLLaf.vtZz6e")))

    # Preferred: fetch values using explicit nearby labels
    indexed = None
    non_indexed = None
    try:
        indexed_node = driver.find_element(
            By.XPATH,
            "//span[contains(@class,'kKiCtc') and normalize-space()='Indexed']"
            "/ancestor::div[contains(@class,'V3oFR')][1]"
            "/following::div[contains(@class,'nnLLaf')][1]"
        )
        iv = normalize_number_string(indexed_node.get_attribute('title') or indexed_node.text or "")
        indexed = iv if iv is not None else None
    except Exception:
        pass
    try:
        non_indexed_node = driver.find_element(
            By.XPATH,
            "//span[contains(@class,'kKiCtc') and translate(normalize-space(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz')='not indexed']"
            "/ancestor::div[contains(@class,'V3oFR')][1]"
            "/following::div[contains(@class,'nnLLaf')][1]"
        )
        niv = normalize_number_string(non_indexed_node.get_attribute('title') or non_indexed_node.text or "")
        non_indexed = niv if niv is not None else None
    except Exception:
        pass

    # Fallback: derive by scanning value tiles with contextual labels
    if indexed is None or non_indexed is None:
        elements = driver.find_elements(By.CSS_SELECTOR, "div.nnLLaf.vtZz6e")
        values_with_context = []
        for el in elements:
            title = el.get_attribute("title") or ""
            text = el.text or ""
            val = normalize_number_string(title) or normalize_number_string(text)
            ctx = ""
            try:
                label_el = el.find_element(
                    By.XPATH,
                    "preceding::div[contains(@class,'V3oFR')][1]"
                )
                ctx = (label_el.text or "").lower()
            except Exception:
                ctx = (text or "").lower()
            values_with_context.append((val, ctx))
        for val, ctx in values_with_context:
            if indexed is None and "indexed" in ctx and "not" not in ctx:
                indexed = val
            if non_indexed is None and "not indexed" in ctx:
                non_indexed = val
        # Last resort: numeric heuristic
        nums_only = [v for v, _ in values_with_context if v and v > 0]
        if (indexed is None or indexed == 0) and nums_only:
            indexed = max(nums_only)
        if (non_indexed is None or non_indexed == 0) and nums_only:
            others = [n for n in nums_only if n != indexed]
            non_indexed = max(others) if others else 0
    indexed = int(indexed or 0)
    non_indexed = int(non_indexed or 0)
    return {"indexed": indexed, "non_indexed": non_indexed}


def _read_primary_metric_numbers(driver) -> tuple[int, int]:
    """Read primary metric numbers from the page cards."""
    cards = driver.find_elements(By.CSS_SELECTOR, "div.nnLLaf.vtZz6e")
    if len(cards) < 3:
        cards = driver.find_elements(By.CSS_SELECTOR, "div[title][class*=nnLLaf]")
    total_requests, avg_response_ms = 0, 0
    if len(cards) >= 1:
        total_requests = normalize_number_string(cards[0].get_attribute("title") or cards[0].text)
    if len(cards) >= 3:
        avg_response_ms = normalize_number_string(cards[2].get_attribute("title") or cards[2].text)
    return total_requests, avg_response_ms


def get_crawl_stats(domain: str, driver, output_dir) -> Dict[str, int]:
    """Fetch crawl statistics from GSC."""
    url = f"https://search.google.com/search-console/settings/crawl-stats?resource_id=sc-domain:{domain}"
    driver.get(url)
    wait = WebDriverWait(driver, 25)
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.nnLLaf.vtZz6e")))
    except TimeoutException:
        pass

    # Ensure "Average response time (ms)" series is enabled: checkmark toggle_2
    try:
        toggle = driver.find_element(By.CSS_SELECTOR, "[data-guidedhelpid='toggle_2']")
        state = (toggle.get_attribute("aria-checked") or toggle.get_attribute("aria-pressed") or "").lower()
        if state not in ("true", "mixed"):
            try:
                toggle.click()
                time.sleep(0.2)
            except Exception:
                pass
    except Exception:
        pass

    # Download Excel file and extract values
    today_total, today_avg_ms = 0, 0
    seven_total, seven_avg_ms = 0, 0
    
    try:
        # Click EXPORT button
        export_btn = None
        try:
            export_btn = driver.find_element(By.CSS_SELECTOR, ".izuYW")
        except Exception:
            pass
        if not export_btn:
            btns = driver.find_elements(By.XPATH, "//*[normalize-space(text())='EXPORT' or normalize-space(text())='Export']")
            if btns:
                export_btn = btns[0]
        
        if export_btn:
            ActionChains(driver).move_to_element(export_btn).click().perform()
            time.sleep(0.5)
            
            # Find and click "Download Excel" option
            excel_option = None
            try:
                candidates = driver.find_elements(By.CSS_SELECTOR, "div.jO7h3c")
                for candidate in candidates:
                    text = (candidate.text or "").strip()
                    if text.lower() == "download excel":
                        excel_option = candidate
                        break
            except Exception:
                pass
            
            if not excel_option:
                for xp in [
                    "//div[@class='jO7h3c' and normalize-space(text())='Download Excel']",
                    "//div[@class='jO7h3c' and contains(normalize-space(text()), 'Download Excel')]",
                    "//*[normalize-space(text())='Download Excel']",
                    "//*[contains(normalize-space(text()), 'Download Excel')]",
                    "//*[contains(normalize-space(text()), 'Excel')]",
                ]:
                    candidates = driver.find_elements(By.XPATH, xp)
                    if candidates:
                        excel_option = candidates[0]
                        break
            
            if excel_option:
                download_url = excel_option.get_attribute('href')
                if download_url:
                    driver.get(download_url)
                else:
                    excel_option.click()
                
                # Wait for download to complete
                time.sleep(5)
                
                # Find the downloaded file
                downloaded_files = []
                for ext in ["*.xlsx", "*.xls", "*.zip"]:
                    downloaded_files.extend(glob.glob(os.path.join(output_dir, ext)))
                
                if not downloaded_files:
                    time.sleep(3)
                    for ext in ["*.xlsx", "*.xls", "*.zip"]:
                        downloaded_files.extend(glob.glob(os.path.join(output_dir, ext)))
                
                if downloaded_files:
                    excel_file = max(downloaded_files, key=os.path.getmtime)
                    
                    # Handle ZIP file (extract if needed)
                    actual_excel_file = excel_file
                    if excel_file.endswith('.zip'):
                        with zipfile.ZipFile(excel_file, 'r') as zip_ref:
                            zip_ref.extractall(output_dir)
                            extracted = []
                            for ext in ["*.xlsx", "*.xls"]:
                                extracted.extend(glob.glob(os.path.join(output_dir, ext)))
                            if extracted:
                                actual_excel_file = max(extracted, key=os.path.getmtime)
                    
                    # Read and parse Excel file
                    try:
                        import warnings
                        import openpyxl
                        with warnings.catch_warnings():
                            warnings.simplefilter("ignore")
                            workbook = openpyxl.load_workbook(actual_excel_file, data_only=True)
                        sheet = workbook.active
                        
                        # Find header row
                        header_row = None
                        header = []
                        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                            row_values = [str(cell).lower() if cell else "" for cell in row]
                            if any("total crawl" in val or "average response" in val for val in row_values):
                                header_row = row_idx
                                header = [str(cell).strip() if cell else "" for cell in row]
                                break
                        
                        if not header_row:
                            header_row = 1
                            header = [str(cell).strip() if cell else "" for cell in sheet[1]]
                        
                        # Find column indices
                        def find_col_idx(search_term: str) -> int:
                            search_term = search_term.lower()
                            for idx, col in enumerate(header):
                                if search_term in (col or "").lower():
                                    return idx
                            return -1
                        
                        idx_total = find_col_idx("total crawl")
                        idx_avg = find_col_idx("average response")
                        
                        # Find the last row with data
                        last_data_row = sheet.max_row
                        for row_idx in range(sheet.max_row, header_row, -1):
                            row = sheet[row_idx]
                            if any(cell.value for cell in row):
                                last_data_row = row_idx
                                break
                        
                        # Extract today's values from the last row
                        if last_data_row > header_row:
                            last_row = sheet[last_data_row]
                            if idx_total >= 0 and idx_total < len(last_row):
                                cell_value = last_row[idx_total].value
                                today_total = normalize_number_string(str(cell_value) if cell_value is not None else "0")
                            if idx_avg >= 0 and idx_avg < len(last_row):
                                cell_value = last_row[idx_avg].value
                                today_avg_ms = normalize_number_string(str(cell_value) if cell_value is not None else "0")
                        
                        # Go up 7 rows from the bottom
                        seven_days_ago_row_num = last_data_row - 7
                        if seven_days_ago_row_num > header_row:
                            seven_row = sheet[seven_days_ago_row_num]
                            if idx_total >= 0 and idx_total < len(seven_row):
                                cell_value = seven_row[idx_total].value
                                seven_total = normalize_number_string(str(cell_value) if cell_value is not None else "0")
                            if idx_avg >= 0 and idx_avg < len(seven_row):
                                cell_value = seven_row[idx_avg].value
                                seven_avg_ms = normalize_number_string(str(cell_value) if cell_value is not None else "0")
                    except ImportError:
                        print("[error] openpyxl not installed. Install it with: pip install openpyxl")
                    except Exception:
                        pass
                    finally:
                        # Cleanup: delete downloaded files
                        try:
                            if 'actual_excel_file' in locals() and actual_excel_file and os.path.exists(actual_excel_file):
                                os.remove(actual_excel_file)
                            if 'excel_file' in locals() and excel_file and os.path.exists(excel_file) and excel_file != actual_excel_file:
                                os.remove(excel_file)
                        except Exception:
                            pass
            
            # Fallback to card values if Excel extraction failed
            if today_total == 0 or today_avg_ms == 0:
                today_total, today_avg_ms = _read_primary_metric_numbers(driver)
    except Exception:
        # Fallback to card values
        today_total, today_avg_ms = _read_primary_metric_numbers(driver)

    return {
        "today_total_requests": today_total,
        "today_avg_response_ms": today_avg_ms,
        "seven_days_total_requests": seven_total,
        "seven_days_avg_response_ms": seven_avg_ms,
    }


def update_sheet_with_results(
    service_account_json: str,
    spreadsheet_id: str,
    target_gid: Optional[int],
    domain: str,
    indexed: int,
    non_indexed: int,
    today_total: int,
    today_avg_ms: int,
    seven_total: int,
    seven_avg_ms: int,
) -> None:
    """
    Update Google Sheet with GSC data results.
    
    Args:
        service_account_json: Path to Google service account JSON file
        spreadsheet_id: Google Sheets spreadsheet ID
        target_gid: Optional sheet tab ID (gid)
        domain: Domain to update
        indexed: Number of indexed URLs
        non_indexed: Number of non-indexed URLs
        today_total: Today's total crawl requests
        today_avg_ms: Today's average response time (ms)
        seven_total: 7 days ago total crawl requests
        seven_avg_ms: 7 days ago average response time (ms)
    """
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ImportError:
        raise ImportError("Google API libraries not installed. Install with: pip install google-api-python-client google-auth")
    
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file(service_account_json, scopes=scopes)
    service = build("sheets", "v4", credentials=creds)
    
    # Resolve sheet title by gid (sheetId) if provided
    sheet_title = None
    if target_gid is not None:
        meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        for sh in meta.get("sheets", []):
            if sh.get("properties", {}).get("sheetId") == target_gid:
                sheet_title = sh.get("properties", {}).get("title")
                break
    if not sheet_title:
        # default to first sheet title
        meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheet_title = meta.get("sheets", [])[0].get("properties", {}).get("title")
    
    # Read first 2 header rows + data
    read_range = f"'{sheet_title}'!A1:ZZ10000"
    resp = service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=read_range).execute()
    rows = resp.get("values", [])
    if not rows:
        return
    
    # Build combined headers from first two rows (handles grouped headers)
    header_row1 = rows[0] if len(rows) >= 1 else []
    header_row2 = rows[1] if len(rows) >= 2 else []
    num_cols = max(len(header_row1), len(header_row2))
    headers = []
    for i in range(num_cols):
        h1 = (header_row1[i] if i < len(header_row1) else "").strip()
        h2 = (header_row2[i] if i < len(header_row2) else "").strip()
        combined = " ".join([p for p in [h1, h2] if p]).strip()
        headers.append(combined or h1 or h2 or f"Col{i+1}")
    
    # Find important column indices
    def find_col(predicates):
        for idx, h in enumerate(headers):
            hl = h.lower()
            if all(p in hl for p in predicates):
                return idx
        return -1
    
    def find_col_any(groups: list[list[str]]) -> int:
        for preds in groups:
            idx = find_col(preds)
            if idx != -1:
                return idx
        return -1
    
    col_domain = find_col(["domain"])
    # Debug: print detected columns
    print(f"🔍 Detected columns - Domain: {col_domain}, Headers: {headers[:10]}")
    
    # If domain not found, try to find it by checking actual data rows
    if col_domain == -1:
        # Look for column that contains domain-like values (has dots, like ubuy.sc)
        for test_row_idx in range(2, min(10, len(rows))):
            row = rows[test_row_idx]
            for col_idx in range(min(len(row), len(headers))):
                cell_value = str(row[col_idx]).strip().lower()
                if "." in cell_value and ("ubuy" in cell_value or "com" in cell_value or len(cell_value.split(".")) >= 2):
                    # This looks like a domain column
                    col_domain = col_idx
                    print(f"🔍 Found domain column by pattern matching: column {col_idx} (header: '{headers[col_idx] if col_idx < len(headers) else 'N/A'}')")
                    break
            if col_domain != -1:
                break
    
    col_last_update = find_col(["last", "update", "date"])
    col_indexed = find_col(["indexed", "urls"])
    
    # Prefer a distinct "no-indexed" column
    col_noindexed = -1
    for idx, h in enumerate(headers):
        hl = h.lower()
        if "no-indexed" in hl or "no indexed" in hl or "non-indexed" in hl or "non indexed" in hl:
            col_noindexed = idx
            break
    
    # Attempt disambiguation using second header row (subheaders)
    if (col_indexed == -1 or col_noindexed == -1) and header_row2:
        second_headers = [(header_row2[i] if i < len(header_row2) else "").strip().lower() for i in range(num_cols)]
        cand_indexed = []
        cand_noindexed = []
        for i, h2 in enumerate(second_headers):
            if not h2:
                continue
            if ("no" in h2 and "index" in h2) or ("non" in h2 and "index" in h2):
                cand_noindexed.append(i)
            elif "indexed" in h2 and "url" in h2:
                cand_indexed.append(i)
        if col_indexed == -1 and cand_indexed:
            col_indexed = cand_indexed[0]
        if col_noindexed == -1 and cand_noindexed:
            col_noindexed = cand_noindexed[0]
    
    # Final fallbacks for two-column group
    if col_noindexed == -1 and col_indexed != -1:
        col_noindexed = col_indexed + 1
    if col_indexed == -1 and col_noindexed != -1:
        col_indexed = max(0, col_noindexed - 1)
    
    # Previous 7 day totals/avg
    col_prev_total = find_col_any([
        ["previous", "total"],
        ["previous", "requests"],
        ["crawl", "previous", "total"],
        ["crawl", "previous", "requests"],
    ])
    col_prev_avg = find_col_any([
        ["previous", "avg"],
        ["previous", "average"],
        ["previous", "response"],
        ["crawl", "previous", "avg"],
        ["crawl", "previous", "response"],
    ])
    if col_prev_avg == -1 and col_prev_total != -1:
        col_prev_avg = col_prev_total + 1
    
    # Latest update totals/avg
    col_latest_total = find_col_any([
        ["latest", "total"],
        ["latest", "requests"],
        ["crawl", "latest", "total"],
        ["crawl", "latest", "requests"],
        ["latest", "update", "total"],
        ["latest", "update", "requests"],
    ])
    col_latest_avg = find_col_any([
        ["latest", "avg"],
        ["latest", "average"],
        ["latest", "response"],
        ["crawl", "latest", "avg"],
        ["crawl", "latest", "response"],
        ["latest", "update", "avg"],
        ["latest", "update", "response"],
    ])
    if col_latest_avg == -1 and col_latest_total != -1:
        col_latest_avg = col_latest_total + 1
    
    # Locate row by domain (case-insensitive match)
    if col_domain == -1:
        print(f"❌ Could not find domain column in sheet. Available headers: {headers[:15]}")
        return
    
    target_row_idx = -1
    domain_lower = domain.lower()
    print(f"🔍 Searching for domain: '{domain}' (normalized: '{domain_lower}')")
    
    for i in range(2, len(rows)):  # skip first two header rows
        row = rows[i]
        if col_domain >= len(row):
            continue
        cell = str(row[col_domain]).strip().lower()
        # Try exact match first
        if cell == domain_lower:
            target_row_idx = i
            print(f"✅ Found exact match at row {i+1}: '{cell}'")
            break
        # Try without protocol if present
        if cell.replace("https://", "").replace("http://", "").replace("www.", "") == domain_lower.replace("www.", ""):
            target_row_idx = i
            print(f"✅ Found match (after cleanup) at row {i+1}: '{cell}'")
            break
    
    if target_row_idx == -1:
        # Try contains match
        for i in range(2, len(rows)):
            row = rows[i]
            if col_domain >= len(row):
                continue
            cell = str(row[col_domain]).strip().lower()
            if domain_lower in cell or cell in domain_lower:
                target_row_idx = i
                print(f"✅ Found partial match at row {i+1}: '{cell}'")
                break
    
    if target_row_idx == -1:
        # Show sample domains for debugging
        sample_domains = []
        for i in range(2, min(10, len(rows))):
            row = rows[i]
            if col_domain < len(row):
                sample_domains.append(str(row[col_domain]).strip())
        print(f"❌ Domain '{domain}' not found in sheet.")
        print(f"   Sample domains in column {col_domain}: {sample_domains[:5]}")
        print(f"   Please check that the domain matches exactly (case-insensitive)")
        return
    
    # Ensure row has enough columns
    while len(rows[target_row_idx]) < len(headers):
        rows[target_row_idx].append("")
    
    # Update fields
    today_str = datetime.now().strftime("%d %b %Y")
    updates_made = []
    
    if col_last_update >= 0:
        rows[target_row_idx][col_last_update] = today_str
        updates_made.append(f"Last Update Date -> {today_str}")
    if col_indexed >= 0:
        rows[target_row_idx][col_indexed] = str(indexed)
        updates_made.append(f"Indexed URLs -> {indexed}")
    if col_noindexed >= 0:
        rows[target_row_idx][col_noindexed] = str(non_indexed)
        updates_made.append(f"No-Indexed URLs -> {non_indexed}")
    
    # Map values to column groups in sheet order
    if col_prev_total >= 0:
        rows[target_row_idx][col_prev_total] = str(int(seven_total))
        updates_made.append(f"Previous 7 Day Total -> {seven_total}")
    if col_prev_avg >= 0:
        rows[target_row_idx][col_prev_avg] = str(int(seven_avg_ms))
        updates_made.append(f"Previous 7 Day Avg -> {seven_avg_ms}")
    if col_latest_total >= 0:
        rows[target_row_idx][col_latest_total] = str(int(today_total))
        updates_made.append(f"Latest Total -> {today_total}")
    if col_latest_avg >= 0:
        rows[target_row_idx][col_latest_avg] = str(int(today_avg_ms))
        updates_made.append(f"Latest Avg -> {today_avg_ms}")
    
    print(f"📝 Updates to make: {updates_made}")
    
    # Write back only that row
    write_start_col = "A"
    write_start_row = target_row_idx + 1  # 1-based
    write_end_col_index = len(headers) - 1
    
    # Convert end column index to Excel letter(s)
    def col_letter(idx_zero_based: int) -> str:
        s = ""
        n = idx_zero_based + 1
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s
    
    write_end_col = col_letter(write_end_col_index)
    write_range = f"'{sheet_title}'!{write_start_col}{write_start_row}:{write_end_col}{write_start_row}"
    body = {"values": [rows[target_row_idx][: len(headers)]]}
    
    try:
        result = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=write_range,
            valueInputOption="USER_ENTERED",
            body=body,
        ).execute()
        updated_cells = result.get('updatedCells', 0)
        print(f"✅ Sheet updated successfully for domain: {domain}")
        print(f"   Updated {updated_cells} cells in range: {write_range}")
    except Exception as e:
        print(f"❌ Error updating sheet: {e}")
        raise


def get_gsc_data(
    domain: str, 
    driver,output_dir
) -> Dict[str, int]:
    """
    Main function to fetch Google Search Console data.
    
    Args:
        domain: Domain property (e.g., "example.com")
        cookies_path: Path to cookies.json file
        headless: Run browser in headless mode (default: True)
        update_sheet: Whether to update Google Sheet (default: False)
        service_account_json: Path to Google service account JSON file (required if update_sheet=True)
        spreadsheet_id: Google Sheets spreadsheet ID (required if update_sheet=True)
        sheet_gid: Optional sheet tab ID (gid), defaults to first sheet
    
    Returns:
        Dictionary containing:
        - indexed: Number of indexed URLs
        - non_indexed: Number of non-indexed URLs
        - today_total_requests: Total crawl requests today
        - today_avg_response_ms: Average response time today (ms)
        - seven_days_total_requests: Total crawl requests 7 days ago
        - seven_days_avg_response_ms: Average response time 7 days ago (ms)
    
    Raises:
        FileNotFoundError: If cookies_path doesn't exist
        Exception: For other errors during execution
    """
    
    try:
        index_data = get_index_counts(domain, driver)
    except TimeoutException:
        print("❌ Could not find the metrics on the page. Check login or cookies.")
        index_data = {"indexed": 0, "non_indexed": 0}
    except Exception as e:
        print(f"❌ Error fetching index counts: {e}")
        index_data = {"indexed": 0, "non_indexed": 0}
        
        # Get crawl stats
    try:
        crawl_data = get_crawl_stats(domain, driver, output_dir)
    except Exception as e:
        print(f"❌ Error fetching crawl stats: {e}")
        crawl_data = {
            "today_total_requests": 0,
            "today_avg_response_ms": 0,
            "seven_days_total_requests": 0,
            "seven_days_avg_response_ms": 0,
        }
        
        # Combine results
    results = {
        "indexed": index_data.get("indexed", 0),
        "non_indexed": index_data.get("non_indexed", 0),
        "today_total_requests": crawl_data.get("today_total_requests", 0),
        "today_avg_response_ms": crawl_data.get("today_avg_response_ms", 0),
        "seven_days_total_requests": crawl_data.get("seven_days_total_requests", 0),
        "seven_days_avg_response_ms": crawl_data.get("seven_days_avg_response_ms", 0),
    }
    service_account_json = get_service_account_path()

    spreadsheet_id = "1fnfgAWfkMzpY4bXWvWdrS0hq5OmoQd_9Ks99y7XrFbo"
    sheet_gid = 0
    if os.path.exists(service_account_json):
            try:
                update_sheet_with_results(
                    service_account_json=service_account_json,
                    spreadsheet_id=spreadsheet_id,
                    target_gid=sheet_gid,
                    domain=domain,
                    indexed=results["indexed"],
                    non_indexed=results["non_indexed"],
                    today_total=results["today_total_requests"],
                    today_avg_ms=results["today_avg_response_ms"],
                    seven_total=results["seven_days_total_requests"],
                    seven_avg_ms=results["seven_days_avg_response_ms"],
                )
            except Exception as e:
                    print(f"⚠️  Failed to update sheet: {e}")
    else:
            print(f"⚠️  Service account file not found: {service_account_json}")
        
    return results
    

